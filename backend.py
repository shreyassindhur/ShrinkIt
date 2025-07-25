import os
import zlib
import json
from flask import Flask, request, send_from_directory, jsonify
from flask_cors import CORS
from werkzeug.utils import secure_filename

# Import libraries for actual file operations
from PIL import Image, ImageFilter, ImageEnhance # For image compression, conversions, and filters
import fitz # PyMuPDF for PDF operations (e.g., PDF to image)
import openpyxl # For XLSX operations
import csv # For CSV operations
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas as rl_canvas # Renamed to avoid conflict with Flask's canvas
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors

# Initialize Flask app
app = Flask(__name__)
# Enable CORS for all origins, allowing frontend to communicate
CORS(app)

# Configuration for file uploads
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # Increased to 500 MB for larger files

# Allowed file extensions for general upload and processing
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'xlsx', 'csv', 'webp'}
IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'} # Helper set for image ops

def allowed_file(filename):
    """Checks if the file extension is allowed."""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_size_to_bytes(size_str):
    """
    Parses a size string (e.g., '10KB', '2MB', '5GB', '50%') into bytes.
    Returns None if empty or invalid format.
    Returns a dict {'percentage': float} if percentage.
    """
    if not size_str: # Handle empty string for automatic compression
        return None

    size_str = size_str.strip().upper()
    if size_str.endswith('%'):
        try:
            percentage = float(size_str[:-1])
            if not (0 <= percentage <= 100):
                return None # Invalid percentage
            return {'percentage': percentage / 100.0}
        except ValueError:
            return None
    elif size_str.endswith('KB'):
        return int(float(size_str[:-2]) * 1024)
    elif size_str.endswith('MB'):
        return int(float(size_str[:-2]) * 1024 * 1024)
    elif size_str.endswith('GB'):
        return int(float(size_str[:-2]) * 1024 * 1024 * 1024)
    elif size_str.isdigit(): # If only a number, assume bytes
        return int(size_str)
    return None # Invalid format

# --- Helper Functions for File Operations ---

def _compress_image(filepath, output_filepath, desired_size_info):
    """
    Compresses an image using Pillow.
    If desired_size_info is a percentage, applies quality reduction.
    Otherwise, applies a default quality.
    """
    try:
        img = Image.open(filepath)
        img_format = img.format # Preserve original format if possible

        # Default quality if no desired_size_info or percentage is provided
        quality = 85
        if isinstance(desired_size_info, dict) and 'percentage' in desired_size_info:
            quality = int(desired_size_info['percentage'] * 100)
            if quality < 10: quality = 10 # Minimum quality to avoid extreme degradation
        
        img.save(output_filepath, img_format, quality=quality, optimize=True)
        return True
    except Exception as e:
        print(f"Error compressing image: {e}")
        return False

def _compress_general(filepath, output_filepath):
    """Compresses general file types using zlib."""
    try:
        with open(filepath, 'rb') as f_in:
            data = f_in.read()
            # zlib.compress uses Z_DEFAULT_COMPRESSION by default (level 6)
            compressed_data = zlib.compress(data)
        with open(output_filepath, 'wb') as f_out:
            f_out.write(compressed_data)
        return True
    except Exception as e:
        print(f"Error compressing general file: {e}")
        return False

def _convert_image_to_pdf(filepath, output_filepath):
    """Converts an image to PDF using Pillow."""
    try:
        img = Image.open(filepath)
        if img.mode == 'RGBA': # Convert RGBA to RGB for PDF compatibility
            img = img.convert('RGB')
        img.save(output_filepath, "PDF", resolution=100.0)
        return True
    except Exception as e:
        print(f"Error converting image to PDF: {e}")
        return False

def _convert_pdf_to_image(filepath, output_filepath_base, output_format):
    """Converts PDF pages to images using PyMuPDF."""
    try:
        doc = fitz.open(filepath)
        # For simplicity, convert only the first page
        page = doc.load_page(0) # Load first page
        pix = page.get_pixmap()
        output_image_path = f"{output_filepath_base}.{output_format}"
        pix.save(output_image_path)
        doc.close()
        return output_image_path # Return the actual path of the generated image
    except Exception as e:
        print(f"Error converting PDF to image: {e}")
        return False

def _convert_image_to_image(filepath, output_filepath, output_format):
    """Converts an image to another image format using Pillow."""
    try:
        img = Image.open(filepath)
        img.save(output_filepath, output_format.upper())
        return True
    except Exception as e:
        print(f"Error converting image to image: {e}")
        return False

def _convert_txt_to_pdf(filepath, output_filepath):
    """Converts a TXT file to PDF using ReportLab."""
    try:
        c = rl_canvas.Canvas(output_filepath, pagesize=letter)
        textobject = c.beginText()
        textobject.setTextOrigin(50, 750) # Start position
        textobject.setFont("Helvetica", 12)

        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                textobject.textLine(line.strip())
        c.drawText(textobject)
        c.save()
        return True
    except Exception as e:
        print(f"Error converting TXT to PDF: {e}")
        return False

def _convert_xlsx_to_csv(filepath, output_filepath):
    """Converts an XLSX file to CSV using openpyxl and csv."""
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active # Get the active sheet

        with open(output_filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in sheet.iter_rows():
                writer.writerow([cell.value for cell in row])
        return True
    except Exception as e:
        print(f"Error converting XLSX to CSV: {e}")
        return False

def _convert_csv_to_xlsx(filepath, output_filepath):
    """Converts a CSV file to XLSX using csv and openpyxl."""
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            for row in reader:
                sheet.append(row)
        workbook.save(output_filepath)
        return True
    except Exception as e:
        print(f"Error converting CSV to XLSX: {e}")
        return False

def _convert_csv_to_pdf(filepath, output_filepath):
    """Converts a CSV file to PDF using ReportLab, rendering as a table."""
    try:
        doc = SimpleDocTemplate(output_filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        data = []
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            for row in reader:
                data.append(row)

        if not data:
            story.append(Paragraph("No data found in CSV file.", styles['Normal']))
        else:
            # Create a table
            table = Table(data)

            # Add style to the table
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey), # Header background
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), # Header text color
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'), # All text left aligned
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), # Header font
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12), # Header padding
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige), # Row background
                ('GRID', (0, 0), (-1, -1), 1, colors.black), # Grid lines
            ]))
            story.append(table)

        doc.build(story)
        return True
    except Exception as e:
        print(f"Error converting CSV to PDF: {e}")
        return False

# --- Image Editing Helper Functions ---

def _resize_image(filepath, output_filepath, width, height):
    """Resizes an image using Pillow."""
    try:
        img = Image.open(filepath)
        resized_img = img.resize((width, height))
        resized_img.save(output_filepath)
        return True
    except Exception as e:
        print(f"Error resizing image: {e}")
        return False

def _crop_image(filepath, output_filepath, left, top, right, bottom):
    """Crops an image using Pillow."""
    try:
        img = Image.open(filepath)
        cropped_img = img.crop((left, top, right, bottom))
        cropped_img.save(output_filepath)
        return True
    except Exception as e:
        print(f"Error cropping image: {e}")
        return False

def _rotate_image(filepath, output_filepath, degrees):
    """Rotates an image using Pillow."""
    try:
        img = Image.open(filepath)
        rotated_img = img.rotate(degrees, expand=True) # expand=True to avoid cropping
        rotated_img.save(output_filepath)
        return True
    except Exception as e:
        print(f"Error rotating image: {e}")
        return False

def _flip_image(filepath, output_filepath, direction):
    """Flips an image using Pillow."""
    try:
        img = Image.open(filepath)
        if direction == 'horizontal':
            flipped_img = img.transpose(Image.FLIP_LEFT_RIGHT)
        elif direction == 'vertical':
            flipped_img = img.transpose(Image.FLIP_TOP_BOTTOM)
        else:
            return False # Invalid direction

        flipped_img.save(output_filepath)
        return True
    except Exception as e:
        print(f"Error flipping image: {e}")
        return False

def _apply_grayscale(filepath, output_filepath):
    """Applies grayscale filter to an image."""
    try:
        img = Image.open(filepath)
        grayscale_img = img.convert('L') # 'L' mode for grayscale
        grayscale_img.save(output_filepath)
        return True
    except Exception as e:
        print(f"Error applying grayscale: {e}")
        return False

def _apply_sepia(filepath, output_filepath):
    """Applies a sepia filter to an image."""
    try:
        img = Image.open(filepath)
        sepia_img = Image.new('RGB', img.size)
        pixels = img.load()
        sepia_pixels = sepia_img.load()

        for y in range(img.size[1]):
            for x in range(img.size[0]):
                r, g, b = pixels[x, y][:3] # Take only RGB, ignore alpha if present
                # Sepia formula (approximate)
                tr = int(0.393 * r + 0.769 * g + 0.189 * b)
                tg = int(0.349 * r + 0.686 * g + 0.168 * b)
                tb = int(0.272 * r + 0.534 * g + 0.131 * b)

                sepia_pixels[x, y] = (min(255, tr), min(255, tg), min(255, tb))
        
        sepia_img.save(output_filepath)
        return True
    except Exception as e:
        print(f"Error applying sepia filter: {e}")
        return False

def _apply_blur(filepath, output_filepath):
    """Applies a blur filter to an image."""
    try:
        img = Image.open(filepath)
        blurred_img = img.filter(ImageFilter.BLUR)
        blurred_img.save(output_filepath)
        return True
    except Exception as e:
        print(f"Error applying blur filter: {e}")
        return False

def _apply_sharpen(filepath, output_filepath):
    """Applies a sharpen filter to an image."""
    try:
        img = Image.open(filepath)
        sharpened_img = img.filter(ImageFilter.SHARPEN)
        sharpened_img.save(output_filepath)
        return True
    except Exception as e:
        print(f"Error applying sharpen filter: {e}")
        return False


# --- Flask Routes ---

@app.route('/')
def index():
    """Basic route to confirm the server is running."""
    return "File Utility Backend is running!"

@app.route('/upload', methods=['POST'])
def upload_file():
    """
    Handles file uploads from the frontend.
    Saves the file to the UPLOAD_FOLDER.
    """
    if 'file' not in request.files:
        return jsonify({'message': 'No file part in the request'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No selected file'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        return jsonify({
            'message': 'File uploaded successfully',
            'filename': filename,
            'filepath': filepath
        }), 200
    return jsonify({'message': 'File type not allowed'}), 400

@app.route('/compress', methods=['POST'])
def compress_file():
    """
    Compresses a file based on its type and desired size.
    If desired_size is not provided, applies a default compression.
    """
    data = request.get_json()
    filename = data.get('filename')
    # desired_size_str can now be None or empty string
    desired_size_str = data.get('desired_size')

    if not filename:
        return jsonify({'message': 'Filename is required for compression'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found for compression'}), 404

    file_extension = filename.rsplit('.', 1)[1].lower()
    processed_filename = f"compressed_{filename}"
    output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], processed_filename)

    success = False
    message = "Compression failed."

    if file_extension in IMAGE_EXTENSIONS:
        # If desired_size_str is empty, parse_size_to_bytes will return None,
        # which _compress_image will interpret as default quality.
        desired_size_info = parse_size_to_bytes(desired_size_str)
        
        # Check if a specific size/percentage was provided and was invalid
        if desired_size_str and desired_size_info is None and not desired_size_str.endswith('%'):
            message = "Invalid desired size format for image compression. Use e.g., '80%' or leave blank for automatic."
        else:
            success = _compress_image(filepath, output_filepath, desired_size_info)
            if success:
                if desired_size_str:
                    message = f'Image "{filename}" compressed to {desired_size_str} successfully.'
                else:
                    message = f'Image "{filename}" compressed automatically (default quality).'
            else:
                message = f'Failed to compress image "{filename}".'
    else:
        # For other file types, apply general zlib compression (default level 6)
        success = _compress_general(filepath, output_filepath)
        if success:
            original_size = os.path.getsize(filepath)
            compressed_size = os.path.getsize(output_filepath)
            message = f'File "{filename}" compressed from {original_size} bytes to {compressed_size} bytes (zlib).'
        else:
            message = f'Failed to compress file "{filename}" with zlib.'

    if success:
        return jsonify({
            'message': message,
            'processed_filename': processed_filename
        }), 200
    else:
        return jsonify({'message': message}), 500


# Removed the /increase_size route as per user request
# @app.route('/increase_size', methods=['POST'])
# def increase_file_size():
#     """
#     Increases a file's size to a desired target by appending null bytes.
#     """
#     data = request.get_json()
#     filename = data.get('filename')
#     desired_size_str = data.get('desired_size')

#     if not filename or not desired_size_str:
#         return jsonify({'message': 'Filename and desired size are required'}), 400

#     filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
#     if not os.path.exists(filepath):
#         return jsonify({'message': 'File not found'}), 404

#     desired_bytes = parse_size_to_bytes(desired_size_str)
#     if desired_bytes is None or isinstance(desired_bytes, dict): # Exclude percentage for increase size
#         return jsonify({'message': 'Invalid desired size format. Use KB, MB, GB suffixes or just bytes.'}), 400

#     try:
#         current_size = os.path.getsize(filepath)
#         if current_size >= desired_bytes:
#             return jsonify({'message': f'File is already {current_size} bytes, which is greater than or equal to desired size of {desired_bytes} bytes.'}), 200

#         bytes_to_add = desired_bytes - current_size
        
#         # Append null bytes to increase file size
#         with open(filepath, 'ab') as f: # 'ab' opens in append binary mode
#             # Write in chunks to avoid memory issues with very large additions
#             chunk_size = 1024 * 1024 # 1 MB
#             for _ in range(0, bytes_to_add, chunk_size):
#                 f.write(b'\0' * min(chunk_size, bytes_to_add - _))
        
#         new_size = os.path.getsize(filepath)
#         return jsonify({
#             'message': f'File "{filename}" size increased to {new_size} bytes (desired: {desired_bytes} bytes).',
#             'processed_filename': filename # Filename remains the same
#         }), 200

#     except Exception as e:
#         return jsonify({'message': f'Error increasing file size: {str(e)}'}), 500


@app.route('/convert', methods=['POST'])
def convert_file():
    """
    Converts a file from one type to another.
    """
    data = request.get_json()
    filename = data.get('filename')
    output_type = data.get('output_type').lower() # Ensure lowercase for consistency

    if not filename or not output_type:
        return jsonify({'message': 'Filename and output type are required for conversion'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found for conversion'}), 404

    input_extension = filename.rsplit('.', 1)[1].lower()
    base_filename = os.path.splitext(filename)[0]
    processed_filename = f"{base_filename}.{output_type}"
    output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], processed_filename)

    success = False
    message = "Conversion failed."

    # Image Conversions (JPG, PNG, WEBP, GIF)
    if input_extension in IMAGE_EXTENSIONS:
        if output_type == 'pdf':
            success = _convert_image_to_pdf(filepath, output_filepath)
            message = f'Image "{filename}" converted to PDF.'
        elif output_type in IMAGE_EXTENSIONS:
            success = _convert_image_to_image(filepath, output_filepath, output_type)
            message = f'Image "{filename}" converted to {output_type.upper()}.'
        else:
            message = f"Unsupported conversion from {input_extension.upper()} to {output_type.upper()}."
    # PDF Conversions
    elif input_extension == 'pdf':
        if output_type in {'png', 'jpg', 'jpeg'}:
            # _convert_pdf_to_image returns the actual path if successful
            temp_output_filename = f"{base_filename}_page1.{output_type}"
            temp_output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], temp_output_filename)
            actual_output_path = _convert_pdf_to_image(filepath, os.path.join(app.config['UPLOAD_FOLDER'], base_filename + "_page1"), output_type)
            
            if actual_output_path:
                # Rename the processed filename to match the actual output from _convert_pdf_to_image
                processed_filename = os.path.basename(actual_output_path)
                message = f'PDF "{filename}" converted to {output_type.upper()} (first page).'
                success = True
            else:
                message = f'Failed to convert PDF "{filename}" to {output_type.upper()}.'
        else:
            message = f"Unsupported conversion from PDF to {output_type.upper()}."
    # Text Conversions
    elif input_extension == 'txt':
        if output_type == 'pdf':
            success = _convert_txt_to_pdf(filepath, output_filepath)
            message = f'TXT "{filename}" converted to PDF.'
        else:
            message = f"Unsupported conversion from TXT to {output_type.upper()}."
    # XLSX Conversions
    elif input_extension == 'xlsx':
        if output_type == 'csv':
            success = _convert_xlsx_to_csv(filepath, output_filepath)
            message = f'XLSX "{filename}" converted to CSV.'
        else:
            message = f"Unsupported conversion from XLSX to {output_type.upper()}."
    # CSV Conversions
    elif input_extension == 'csv':
        if output_type == 'xlsx':
            success = _convert_csv_to_xlsx(filepath, output_filepath)
            message = f'CSV "{filename}" converted to XLSX.'
        elif output_type == 'pdf': # NEW: CSV to PDF
            success = _convert_csv_to_pdf(filepath, output_filepath)
            message = f'CSV "{filename}" converted to PDF.'
        else:
            message = f"Unsupported conversion from CSV to {output_type.upper()}."
    else:
        message = f"Unsupported input file type: {input_extension.upper()}."

    if success:
        return jsonify({
            'message': message,
            'processed_filename': processed_filename
        }), 200
    else:
        return jsonify({'message': message}), 500

# --- New Image Editing Routes ---

@app.route('/image_resize', methods=['POST'])
def image_resize():
    data = request.get_json()
    filename = data.get('filename')
    width = data.get('width')
    height = data.get('height')

    if not filename or width is None or height is None:
        return jsonify({'message': 'Filename, width, and height are required.'}), 400
    if not isinstance(width, int) or not isinstance(height, int) or width <= 0 or height <= 0:
        return jsonify({'message': 'Width and height must be positive integers.'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found.'}), 404
    if filename.rsplit('.', 1)[1].lower() not in IMAGE_EXTENSIONS:
        return jsonify({'message': 'File is not an image type for resizing.'}), 400

    base_filename, ext = os.path.splitext(filename)
    processed_filename = f"{base_filename}_resized{ext}"
    output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], processed_filename)

    if _resize_image(filepath, output_filepath, width, height):
        return jsonify({
            'message': f'Image "{filename}" resized to {width}x{height}.',
            'processed_filename': processed_filename
        }), 200
    else:
        return jsonify({'message': f'Failed to resize image "{filename}".'}), 500

@app.route('/image_crop', methods=['POST'])
def image_crop():
    data = request.get_json()
    filename = data.get('filename')
    left = data.get('left')
    top = data.get('top')
    right = data.get('right')
    bottom = data.get('bottom')

    if not filename or any(val is None for val in [left, top, right, bottom]):
        return jsonify({'message': 'Filename and crop coordinates (left, top, right, bottom) are required.'}), 400
    if not all(isinstance(val, int) and val >= 0 for val in [left, top, right, bottom]):
        return jsonify({'message': 'Crop coordinates must be non-negative integers.'}), 400
    if left >= right or top >= bottom:
        return jsonify({'message': 'Invalid crop dimensions: right must be > left, bottom must be > top.'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found.'}), 404
    if filename.rsplit('.', 1)[1].lower() not in IMAGE_EXTENSIONS:
        return jsonify({'message': 'File is not an image type for cropping.'}), 400

    base_filename, ext = os.path.splitext(filename)
    processed_filename = f"{base_filename}_cropped{ext}"
    output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], processed_filename)

    if _crop_image(filepath, output_filepath, left, top, right, bottom):
        return jsonify({
            'message': f'Image "{filename}" cropped to ({left},{top},{right},{bottom}).',
            'processed_filename': processed_filename
        }), 200
    else:
        return jsonify({'message': f'Failed to crop image "{filename}".'}), 500

@app.route('/image_rotate', methods=['POST'])
def image_rotate():
    data = request.get_json()
    filename = data.get('filename')
    degrees = data.get('degrees')

    if not filename or degrees is None:
        return jsonify({'message': 'Filename and degrees are required.'}), 400
    if not isinstance(degrees, (int, float)):
        return jsonify({'message': 'Degrees must be a number.'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found.'}), 404
    if filename.rsplit('.', 1)[1].lower() not in IMAGE_EXTENSIONS:
        return jsonify({'message': 'File is not an image type for rotation.'}), 400

    base_filename, ext = os.path.splitext(filename)
    processed_filename = f"{base_filename}_rotated{ext}"
    output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], processed_filename)

    if _rotate_image(filepath, output_filepath, degrees):
        return jsonify({
            'message': f'Image "{filename}" rotated by {degrees} degrees.',
            'processed_filename': processed_filename
        }), 200
    else:
        return jsonify({'message': f'Failed to rotate image "{filename}".'}), 500

@app.route('/image_flip', methods=['POST'])
def image_flip():
    data = request.get_json()
    filename = data.get('filename')
    direction = data.get('direction') # 'horizontal' or 'vertical'

    if not filename or direction is None:
        return jsonify({'message': 'Filename and flip direction are required.'}), 400
    if direction not in ['horizontal', 'vertical']:
        return jsonify({'message': 'Direction must be "horizontal" or "vertical".'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found.'}), 404
    if filename.rsplit('.', 1)[1].lower() not in IMAGE_EXTENSIONS:
        return jsonify({'message': 'File is not an image type for flipping.'}), 400

    base_filename, ext = os.path.splitext(filename)
    processed_filename = f"{base_filename}_flipped_{direction}{ext}"
    output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], processed_filename)

    if _flip_image(filepath, output_filepath, direction):
        return jsonify({
            'message': f'Image "{filename}" flipped {direction}.',
            'processed_filename': processed_filename
        }), 200
    else:
        return jsonify({'message': f'Failed to flip image "{filename}".'}), 500

@app.route('/image_grayscale', methods=['POST'])
def image_grayscale():
    data = request.get_json()
    filename = data.get('filename')

    if not filename:
        return jsonify({'message': 'Filename is required.'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found.'}), 404
    if filename.rsplit('.', 1)[1].lower() not in IMAGE_EXTENSIONS:
        return jsonify({'message': 'File is not an image type for grayscale.'}), 400

    base_filename, ext = os.path.splitext(filename)
    processed_filename = f"{base_filename}_grayscale{ext}"
    output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], processed_filename)

    if _apply_grayscale(filepath, output_filepath):
        return jsonify({
            'message': f'Image "{filename}" converted to grayscale.',
            'processed_filename': processed_filename
        }), 200
    else:
        return jsonify({'message': f'Failed to apply grayscale to image "{filename}".'}), 500

@app.route('/image_sepia', methods=['POST'])
def image_sepia():
    data = request.get_json()
    filename = data.get('filename')

    if not filename:
        return jsonify({'message': 'Filename is required.'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found.'}), 404
    if filename.rsplit('.', 1)[1].lower() not in IMAGE_EXTENSIONS:
        return jsonify({'message': 'File is not an image type for sepia.'}), 400

    base_filename, ext = os.path.splitext(filename)
    processed_filename = f"{base_filename}_sepia{ext}"
    output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], processed_filename)

    if _apply_sepia(filepath, output_filepath):
        return jsonify({
            'message': f'Sepia filter applied to image "{filename}".',
            'processed_filename': processed_filename
        }), 200
    else:
        return jsonify({'message': f'Failed to apply sepia filter to image "{filename}".'}), 500

@app.route('/image_blur', methods=['POST'])
def image_blur():
    data = request.get_json()
    filename = data.get('filename')

    if not filename:
        return jsonify({'message': 'Filename is required.'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found.'}), 404
    if filename.rsplit('.', 1)[1].lower() not in IMAGE_EXTENSIONS:
        return jsonify({'message': 'File is not an image type for blur.'}), 400

    base_filename, ext = os.path.splitext(filename)
    processed_filename = f"{base_filename}_blurred{ext}"
    output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], processed_filename)

    if _apply_blur(filepath, output_filepath):
        return jsonify({
            'message': f'Blur filter applied to image "{filename}".',
            'processed_filename': processed_filename
        }), 200
    else:
        return jsonify({'message': f'Failed to apply blur filter to image "{filename}".'}), 500

@app.route('/image_sharpen', methods=['POST'])
def image_sharpen():
    data = request.get_json()
    filename = data.get('filename')

    if not filename:
        return jsonify({'message': 'Filename is required.'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found.'}), 404
    if filename.rsplit('.', 1)[1].lower() not in IMAGE_EXTENSIONS:
        return jsonify({'message': 'File is not an image type for sharpen.'}), 400

    base_filename, ext = os.path.splitext(filename)
    processed_filename = f"{base_filename}_sharpened{ext}"
    output_filepath = os.path.join(app.config['UPLOAD_FOLDER'], processed_filename)

    if _apply_sharpen(filepath, output_filepath):
        return jsonify({
            'message': f'Sharpen filter applied to image "{filename}".',
            'processed_filename': processed_filename
        }), 200
    else:
        return jsonify({'message': f'Failed to apply sharpen filter to image "{filename}".'}), 500


@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    """
    Allows downloading of processed files from the UPLOAD_FOLDER.
    """
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(filepath):
        return jsonify({'message': 'File not found for download'}), 404
    
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)
    except Exception as e:
        return jsonify({'message': f'Error serving file for download: {str(e)}'}), 500

@app.route('/delete', methods=['POST'])
def delete_file():
    """
    Deletes a file from the UPLOAD_FOLDER.
    """
    data = request.get_json()
    filename = data.get('filename')

    if not filename:
        return jsonify({'message': 'Filename is required for deletion'}), 400

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        try:
            os.remove(filepath)
            return jsonify({'message': f'File "{filename}" deleted successfully'}), 200
        except Exception as e:
            return jsonify({'message': f'Error deleting file: {str(e)}'}), 500
    else:
        return jsonify({'message': 'File not found for deletion'}), 404

if __name__ == '__main__':
    # Run the Flask app
    # In a production environment, use a production-ready WSGI server like Gunicorn
    app.run(debug=True, port=5000)
